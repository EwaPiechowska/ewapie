

import openpyxl

def tabliczka(rozmiar = 10):
    excel = openpyxl.Workbook()
    excel.save('tabliczka_mnozenia.xlsx')
    arkusze = excel.sheetnames
    arkusz = excel.active
    #wpadamy w 1 wiersz i potem 10 po kolumnach i tak po kolei
    for line in range(1,rozmiar):
        for table in range(1,rozmiar):
            komorka = arkusz.cell(line, table)
            print(line * table, '\n',)
            komorka.value = line * table
    excel.save('tabliczka_mnozenia.xlsx')
tabliczka(40)